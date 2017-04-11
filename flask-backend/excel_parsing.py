import os
import datetime

import xlrd as xlrd
from openpyxl import load_workbook
from openpyxl import styles
import os 
from app.table_model.number.asset_name_model import AssetNameModel
from app.table_model.number.standard_model import StandardModel
from app.table_model.number.buy_model import BuyModel
from app.table_model.number.asset_model import AssetModel
from app.table_model.device.spec.rack_spec_model import RackSpecModel
from app.table_model.device.spec.server_spec_model import ServerSpecModel
from app.table_model.device.spec.storage.storage_spec_model import StorageSpecModel
from app.table_model.device.spec.storage.storage_spec_name_model import StorageSpecNameModel
from app.table_model.device.spec.storage.storage_spec_type_model import StorageSpecTypeModel
from app.table_model.device.spec.switch_spec_model import SwitchSpecModel
from app.table_model.location.detail_location_model import DetailLocationModel
from app.table_model.location.location_model import LocationModel
from app.table_model.device.device_model import DeviceModel
from app.table_model.device.rack_model import RackModel
from app.table_model.device.server_model import ServerModel
from app.table_model.device.storage_model import StorageModel
from app.table_model.device.switch_model import SwitchModel
from app.table_model.rack_location.device_info import DeviceInfo
from app.table_model.rack_location.for_server_model import DeviceInfoForServerModel
from app.table_model.rack_location.for_switch_model import DeviceInfoForSwitchModel
from app.table_model.service.service_model import ServiceModel
from app.table_model.service.service_name_model import ServiceNameModel
from app.table_model.rack_location.rack_location_for_server_model import RackLocationForServerModel
from app.table_model.rack_location.rack_location_for_switch_model import RackLocationForSwitchModel
from app.table_model.rack_location.rack_location_model import RackLocationModel



from sqlalchemy.orm import sessionmaker, scoped_session
from app.orm.session import engine

session_factory = sessionmaker(bind=engine)
Session = scoped_session(session_factory)
session = Session()

ASSET_SHEETS = ['TOTAL', 'SERVER', 'SWITCH', 'STORAGE', 'RACK']

def parsing_asset():
    x = xlrd.open_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), "1_Asset-201703.xlsx"))
    sheet = x.sheet_by_name(ASSET_SHEETS[0])
    assetServerSheet = x.sheet_by_name(ASSET_SHEETS[1])
    assetSwitchSheet = x.sheet_by_name(ASSET_SHEETS[2])
    assetStorageSheet = x.sheet_by_name(ASSET_SHEETS[3])
    assetRackSheet = x.sheet_by_name(ASSET_SHEETS[4])
    service_storageFile = xlrd.open_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), "2_Storage-201703.xlsx"))
    service_storageSheet = service_storageFile.sheet_by_name('Sheet2')
    storageFile = xlrd.open_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), "2_Storage_back.xlsx"))
    storageSheet = storageFile.sheet_by_name('Sheet2')
    ServiceFile = xlrd.open_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), "Service Resources-201703.xlsx"))
    ServiceSheet = ServiceFile.sheet_by_name('2016.10자원현황')
    RackFile = xlrd.open_workbook(os.path.join(os.path.dirname(os.path.abspath(__file__)), "Rack Info-201703.xlsx"))
    RackSheet = RackFile.sheet_by_name('201703')

    #openpyxl
    RackFileOPX = load_workbook("Rack Info-201703.xlsx")
    RackSheetOPX = RackFileOPX["201703"]

    sheetCols = 279
    assetRackSheetCols = 24
    ncol = sheet.ncols
    nlow = sheet.nrows
    
    #""" 
    #####################################
    print("-------- "+ASSET_SHEETS[0]+" --------")
    print("Number of col: " + str(ncol))
    print("Number of low: " + str(nlow))
    
    print("-------- Values of Excel file --------")
    values = set()
    for row in range(1,sheetCols):
        val = sheet.row_values(row)[2]
        values.add(val) 
    print(values)
    for value in values:
        newData = AssetNameModel(asset_name=value)
        session.add(newData)
    session.commit()

    values = set()
    for row in range(2,sheetCols):
        val = sheet.row_values(row)[3]
        values.add(val)
    print(values)
    for value in values:
        newData = StandardModel(standard_name=value)
        session.add(newData)
    session.commit()

    values = set()
    for row in range(2,sheetCols):
        val = sheet.row_values(row)[5]
        values.add(val)
    print(values)
    for value in values:
        newData = BuyModel(buy_name=value)
        session.add(newData)
    session.commit()
    
    for row in range(2,sheetCols):
        asset_num = sheet.row_values(row)[0]
        get_date = sheet.row_values(row)[1]
        asset_name = sheet.row_values(row)[2]
        standard_id = sheet.row_values(row)[3]
        years = sheet.row_values(row)[6]
        price = sheet.row_values(row)[4]
        buy = sheet.row_values(row)[5]
        get_date = datetime.datetime.strptime(get_date, '%Y-%m-%d')
        asset_name_id = session.query(AssetNameModel).filter_by(asset_name=asset_name).first().id
        standard_id = session.query(StandardModel).filter_by(standard_name=standard_id).first().id
        buy = session.query(BuyModel).filter_by(buy_name=buy).first().id
        newData = AssetModel(asset_num=asset_num, get_date=get_date, years=years, price=price)
        newData.asset_name_id = asset_name_id
        newData.standard_id = standard_id
        newData.buy = buy
        session.add(newData)
    session.commit()

    values = set()
    for row in range(6,114):
        val = storageSheet.row_values(row)[0]
        values.add(val)
    values.remove('') 
    print(values)
    for value in values:
        newData = StorageSpecNameModel(spec_name=value)
        session.add(newData)
    session.commit()
    
    values = set()
    for row in range(6,114):
        val = storageSheet.row_values(row)[3]
        values.add(val)
    values.remove('') 
    print(values)
    for value in values:
        newData = StorageSpecTypeModel(spec_type=value)
        session.add(newData)
    session.commit()
    
    for row in range(6,114):
        spec_id = storageSheet.row_values(row)[0]
        if spec_id == '':
            continue
        disk_spec = storageSheet.row_values(row)[2]
        disk_type_id = storageSheet.row_values(row)[3]
        volume = storageSheet.row_values(row)[4]
        spec_id = session.query(StorageSpecNameModel).filter_by(spec_name=spec_id).first().id
        disk_type_id = session.query(StorageSpecTypeModel).filter_by(spec_type=disk_type_id).first().id
        newData = StorageSpecModel(disk_spec=disk_spec, volume=volume)
        newData.spec_id = spec_id
        newData.disk_type_id = disk_type_id
        session.add(newData)
    session.commit()
    
    values = set()
    for row in range(1,assetRackSheetCols):
        val = assetRackSheet.row_values(row)[4]
        values.add(val)
    print(values)
    for value in values:
        newData = RackSpecModel(spec=value)
        session.add(newData)
    session.commit()
    
    values = set()
    for row in range(1,535):
        val = assetServerSheet.row_values(row)[4]
        values.add(val)
    print(values)
    for value in values:
        newData = ServerSpecModel(spec=value)
        session.add(newData)
    session.commit()

    values = set()
    for row in range(1,62):
        val = assetSwitchSheet.row_values(row)[4]
        values.add(val)
    print(values)
    for value in values:
        newData = SwitchSpecModel(spec=value)
        session.add(newData)
    session.commit()

    for row in range(1,assetRackSheetCols):
        asset_id = assetRackSheet.row_values(row)[0]
        manage_num = assetRackSheet.row_values(row)[1]
        spec_id = assetRackSheet.row_values(row)[4]
        rack_size = 42
        asset_id = session.query(AssetModel).filter_by(asset_num=asset_id).first().id
        spec_id = session.query(RackSpecModel).filter_by(spec=spec_id).first().id
        newData = RackModel(asset_id=asset_id, manage_num=manage_num, rack_size=rack_size, spec_id=spec_id)
        session.add(newData)
    session.commit()
    
    values = set()
    for row in range(4,39):
        val = ServiceSheet.row_values(row)[1]
        values.add(val)
    values.remove('') 
    print(values)
    for value in values:
        newData = ServiceNameModel(service_name=value)
        session.add(newData)
    session.commit()

    for row in range(6,114):
        spec_id = service_storageSheet.row_values(row)[0]
        disk_spec = service_storageSheet.row_values(row)[2]
        used_size = service_storageSheet.row_values(row)[6]
        service_name_id = service_storageSheet.row_values(row)[7]
        usage = service_storageSheet.row_values(row)[8]
        spec_id = session.query(StorageSpecNameModel).filter_by(spec_name=spec_id).first().id
        storage_spec_id = session.query(StorageSpecModel).filter_by(spec_id=spec_id).filter_by(disk_spec=disk_spec).first().id
        service_name_id = session.query(ServiceNameModel).filter_by(service_name=service_name_id).first().id
        newData = ServiceModel(storage_spec_id=storage_spec_id, used_size=used_size, service_name_id=service_name_id, usage=usage)
        session.add(newData)
    session.commit()

    values = set()
    for row in range(1,535):
        val = assetServerSheet.row_values(row)[3]
        val = val.split("-")
        values.add(val[0])
    for row in range(1,62):
        val = assetSwitchSheet.row_values(row)[3]
        val = val.split("-")
        values.add(val[0])
    for row in range(1,assetRackSheetCols):
        val = assetRackSheet.row_values(row)[3]
        val = val.split("-")
        values.add(val[0])
    print(values)
    for value in values:
        newData = LocationModel(location=value)
        session.add(newData)
    session.commit()
    
    for row in range(1,assetRackSheetCols):
        manage_num = assetRackSheet.row_values(row)[1]
        val = assetRackSheet.row_values(row)[3]
        location, detail = val.split("-")
        rack_id = session.query(DeviceModel).filter_by(manage_num=manage_num).first().id
        location_id = session.query(LocationModel).filter_by(location=location).first().id
        newData = DetailLocationModel(location_id=location_id, detail=detail, rack_id=rack_id)
        session.add(newData)
    session.commit()

    for row in range(1,535):
        asset_id = assetServerSheet.row_values(row)[0]
        manage_num = assetServerSheet.row_values(row)[1]
        spec_id = assetServerSheet.row_values(row)[4]
        val = assetServerSheet.row_values(row)[3]
        location, detail = val.split("-")
        size = 1
        core_num = assetServerSheet.row_values(row)[5]

        asset_id = session.query(AssetModel).filter_by(asset_num=asset_id).first().id
        spec_id = session.query(ServerSpecModel).filter_by(spec=spec_id).first().id
        location_id = session.query(DetailLocationModel).filter_by(detail=detail).first().id
        newData = ServerModel(asset_id=asset_id, manage_num=manage_num, spec_id=spec_id, location_id=location_id, size=size, core_num=core_num)
        session.add(newData)
    session.commit()

    for row in range(1,62):
        asset_id = assetSwitchSheet.row_values(row)[0]
        manage_num = assetSwitchSheet.row_values(row)[1]
        spec_id = assetSwitchSheet.row_values(row)[4]
        val = assetSwitchSheet.row_values(row)[3]
        location, detail = val.split("-")
        size = 1

        asset_id = session.query(AssetModel).filter_by(asset_num=asset_id).first().id
        spec_id = session.query(SwitchSpecModel).filter_by(spec=spec_id).first().id
        location_id = session.query(DetailLocationModel).filter_by(detail=detail).first().id
        newData = SwitchModel(asset_id=asset_id, manage_num=manage_num, spec_id=spec_id, location_id=location_id, size=size)
        session.add(newData)
    session.commit()

    for col in range(0,23):
        name = RackSheet.row_values(4)[(col*5)+1]
        print(name)
        for row in range(7,49):
            ip_v4 = RackSheet.row_values(row)[(col*5)+1]
            if ip_v4 == '':
                continue
            manage_num = RackSheet.row_values(row)[(col*5)+2]
            if manage_num == '':
                continue
            print(manage_num)
            if manage_num[0] == 'N':
                if session.query(DeviceModel).filter_by(manage_num=manage_num).count() == 0:
                    print(manage_num+"is passed!")
                    continue
                switch_id = session.query(DeviceModel).filter_by(manage_num=manage_num).first().id
                switch_id = session.query(SwitchModel).filter_by(id=switch_id).first().id
                newData = DeviceInfoForSwitchModel(ip_v4=ip_v4, switch_id=switch_id)
                session.add(newData)
                session.commit()
            elif manage_num[0] == 'S':
                if session.query(DeviceModel).filter_by(manage_num=manage_num).count() == 0:
                    print(manage_num+"is passed!")
                    continue
                server_id = session.query(DeviceModel).filter_by(manage_num=manage_num).first().id
                server_id = session.query(ServerModel).filter_by(id=server_id).first().id
                newData = DeviceInfoForServerModel(ip_v4=ip_v4, server_id=server_id)
                session.add(newData)
                session.commit()
            else: 
                print(manage_num)
    #"""
    ServiceList = {}
    for row in range(51, 65):
        color = RackSheetOPX.cell(row=row, column=3).fill.start_color.index
        name = RackSheet.row_values(row-1)[2]
        ServiceList[color]=name

    for col in range(0,24):
        col = (col*5)+3
        rackInfo = RackSheet.row_values(4)[col-2]
        detail, manage_num = rackInfo.split(" - ")
        location_id = session.query(DetailLocationModel).filter_by(detail=detail).first().id
        for row in range(8,49):
            color = RackSheetOPX.cell(row=row, column=col).fill.start_color.index
            manage_num = RackSheetOPX.cell(row=row, column=col).value
            if manage_num is None:
                continue
            print(manage_num)
            if color not in ServiceList:
                print(manage_num + " does not exist service.")
                continue
            if session.query(ServiceNameModel).filter_by(service_name=name).count() == 0:
                print(manage_num + " does not exist service.")
                continue
            name = ServiceList[color]
            print(name)
            service_id = session.query(ServiceNameModel).filter_by(service_name=name).first().id

            start_index = row
            while start_index < 49:
                if (manage_num != RackSheetOPX.cell(row=start_index, column=col).value) or (color != RackSheetOPX.cell(row=start_index, column=col).fill.start_color.index):
                    break
                start_index = start_index + 1
            start_index = (-1 * start_index) + 50
            service_on_off = True
    
            if manage_num[0] == "N":
                if session.query(DeviceModel).filter_by(manage_num=manage_num).count() == 0:
                    print(manage_num+"is passed!")
                    continue
                switch_id = session.query(DeviceModel).filter_by(manage_num=manage_num).first().id
                switch_id = session.query(SwitchModel).filter_by(id=switch_id).first().id
                newData = RackLocationForSwitchModel(start_index=start_index, location_id=location_id, service_id=service_id, service_on_off=service_on_off, switch_id=switch_id)
                session.add(newData)
                session.commit()

            elif manage_num[0] == "S":
                if session.query(DeviceModel).filter_by(manage_num=manage_num).count() == 0:
                    print(manage_num+"is passed!")
                    continue
                server_id = session.query(DeviceModel).filter_by(manage_num=manage_num).first().id
                server_id = session.query(ServerModel).filter_by(id=server_id).first().id
                newData = RackLocationForServerModel(start_index=start_index, location_id=location_id, service_id=service_id, service_on_off=service_on_off, server_id=server_id)
                session.add(newData)
                session.commit()
 
    
    



if __name__ == '__main__':
    parsing_asset()
    pass
